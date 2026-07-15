import io
import os
import unittest
from datetime import date
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import and_, create_engine, func, select
from sqlalchemy.orm import sessionmaker

os.environ.setdefault("JWT_SIGNING_KEYS", '{"dev-test-kid":"dev-test-secret-with-at-least-32-bytes"}')
os.environ.setdefault("JWT_ACTIVE_KID", "dev-test-kid")
os.environ.setdefault("ACCESS_TOKEN_TTL_SECONDS", "1200")
os.environ.setdefault("REFRESH_TOKEN_TTL_SECONDS", "1209600")
os.environ.setdefault("AUTH_COOKIE_SECURE", "false")

from fastapi.testclient import TestClient

from app.auth.passwords import hash_password
from app.auth.tokens import create_access_token
from app.db.models import AppUser, AuditLog, RentConfirm, Room, Site, TenantContract, UserScope
from app.db.session import get_db
from app.main import app
from app.meta.seed import seed_metadata


class DataApiTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        database_url = os.environ.get("DATABASE_URL")
        if not database_url:
            raise unittest.SkipTest("DATABASE_URL is not set")
        cls.engine = create_engine(database_url, future=True)
        cls.SessionLocal = sessionmaker(expire_on_commit=False, future=True)

    @classmethod
    def tearDownClass(cls) -> None:
        cls.engine.dispose()

    def setUp(self) -> None:
        self.connection = self.engine.connect()
        self.transaction = self.connection.begin()
        self.session = self.SessionLocal(bind=self.connection)
        self.prefix = f"data_{uuid4().hex}"
        seed_metadata(self.session, actor="test")
        self.admin = self.create_user("admin")
        self.staff = self.create_user("staff")
        self.site_a = Site(site_code=f"{self.prefix}_A", name="Alpha")
        self.site_b = Site(site_code=f"{self.prefix}_B", name="Beta")
        self.session.add_all([self.site_a, self.site_b])
        self.session.flush()
        self.room_a1 = Room(site_id=self.site_a.id, room_code=f"{self.prefix}_101", room_name="North 101")
        self.room_a2 = Room(site_id=self.site_a.id, room_code=f"{self.prefix}_102", room_name="North 102")
        self.room_b1 = Room(site_id=self.site_b.id, room_code=f"{self.prefix}_201", room_name=None)
        self.session.add_all([self.room_a1, self.room_a2, self.room_b1])
        self.session.flush()
        self.contract_a1 = TenantContract(
            room_id=self.room_a1.id,
            lease_start_date=date(2026, 1, 1),
            lease_end_date=date(2026, 12, 31),
            rent=12000,
            contact_name="Alice",
            contact_phone="0912000001",
        )
        self.contract_a2 = TenantContract(
            room_id=self.room_a2.id,
            lease_start_date=date(2026, 3, 1),
            lease_end_date=None,
            rent=15000,
            contact_name="Bob",
            contact_phone="0912000002",
        )
        self.contract_b1 = TenantContract(
            room_id=self.room_b1.id,
            lease_start_date=date(2025, 11, 1),
            lease_end_date=date(2026, 5, 31),
            rent=18000,
            contact_name="Carol",
            contact_phone="0912000003",
        )
        self.session.add_all([self.contract_a1, self.contract_a2, self.contract_b1])
        self.confirm_a1 = RentConfirm(
            room_id=self.room_a1.id,
            billing_ym="202607",
            charge_type="rent",
            run_version=1,
            status="draft",
            rent_amount=12000,
            electricity_amount=300,
            fixed_fee_amount=500,
            exception_amount=0,
            total_amount=12800,
        )
        self.confirm_a2 = RentConfirm(
            room_id=self.room_a2.id,
            billing_ym="202608",
            charge_type="electricity",
            run_version=2,
            status="published",
            rent_amount=15000,
            electricity_amount=700,
            fixed_fee_amount=500,
            exception_amount=100,
            total_amount=16300,
        )
        self.confirm_b1 = RentConfirm(
            room_id=self.room_b1.id,
            billing_ym="202607",
            charge_type="rent",
            run_version=1,
            status="published",
            rent_amount=18000,
            electricity_amount=900,
            fixed_fee_amount=500,
            exception_amount=None,
            total_amount=19400,
        )
        self.session.add_all([self.confirm_a1, self.confirm_a2, self.confirm_b1])
        self.session.flush()
        self.session.add(UserScope(user_id=self.staff.id, scope_type="site", scope_value=str(self.site_a.id)))
        self.session.commit()

        def override_get_db():
            yield self.session

        app.dependency_overrides[get_db] = override_get_db
        self.client = TestClient(app)

    def tearDown(self) -> None:
        app.dependency_overrides.clear()
        self.session.close()
        self.transaction.rollback()
        self.connection.close()

    def create_user(self, role: str) -> AppUser:
        user = AppUser(
            username=f"{self.prefix}_{role}",
            password_hash=hash_password("password"),
            role=role,
            token_version=0,
        )
        self.session.add(user)
        self.session.flush()
        return user

    def auth_headers(self, user: AppUser) -> dict[str, str]:
        return {"Authorization": f"Bearer {create_access_token(user)}"}

    def query(self, table: str, payload: dict, user: AppUser | None = None):
        return self.client.post(
            f"/api/data/{table}/query", json=payload, headers=self.auth_headers(user or self.admin)
        )

    def test_filter_types_match_direct_db_counts(self) -> None:
        cases = [
            (
                "room",
                [{"col": "room_code", "op": "eq", "val": self.room_a1.room_code}],
                select(func.count()).select_from(Room).where(Room.room_code == self.room_a1.room_code),
            ),
            (
                "room",
                [{"col": "room_name", "op": "contains", "val": "North"}],
                select(func.count()).select_from(Room).where(Room.room_name.ilike("%North%")),
            ),
            (
                "rent_confirm",
                [{"col": "charge_type", "op": "eq", "val": "rent"}],
                select(func.count()).select_from(RentConfirm).where(RentConfirm.charge_type == "rent"),
            ),
            (
                "rent_confirm",
                [{"col": "billing_ym", "op": "range", "val": ["202607", "202607"]}],
                select(func.count())
                .select_from(RentConfirm)
                .where(and_(RentConfirm.billing_ym >= "202607", RentConfirm.billing_ym <= "202607")),
            ),
            (
                "tenant_contract",
                [{"col": "rent", "op": "range", "val": [13000, 20000]}],
                select(func.count())
                .select_from(TenantContract)
                .where(and_(TenantContract.rent >= 13000, TenantContract.rent <= 20000)),
            ),
            (
                "room",
                [{"col": "room_name", "op": "isnull", "val": True}],
                select(func.count()).select_from(Room).where(Room.room_name.is_(None)),
            ),
        ]
        for table, filters, count_statement in cases:
            with self.subTest(table=table, filters=filters):
                response = self.query(table, {"filters": filters, "sort": [], "page": 1, "size": 50})

                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.json()["total"], self.session.scalar(count_statement))

    def test_export_row_count_matches_query_total_and_writes_audit_log(self) -> None:
        payload = {
            "filters": [{"col": "billing_ym", "op": "eq", "val": "202607"}],
            "sort": [{"col": "room_id", "dir": "asc"}],
        }
        query_response = self.query("rent_confirm", {**payload, "page": 1, "size": 50})
        before = self.session.scalar(
            select(func.count())
            .select_from(AuditLog)
            .where(AuditLog.actor == self.admin.id, AuditLog.action == "export")
        )

        export_response = self.client.post(
            "/api/data/rent_confirm/export",
            json=payload,
            headers=self.auth_headers(self.admin),
        )

        self.assertEqual(export_response.status_code, 200)
        workbook = load_workbook(io.BytesIO(export_response.content), read_only=True)
        sheet = workbook.active
        self.assertEqual(sheet.max_row - 1, query_response.json()["total"])
        after = self.session.scalar(
            select(func.count())
            .select_from(AuditLog)
            .where(AuditLog.actor == self.admin.id, AuditLog.action == "export")
        )
        self.assertEqual(after, before + 1)
        audit = self.session.scalar(
            select(AuditLog)
            .where(AuditLog.actor == self.admin.id, AuditLog.action == "export")
            .order_by(AuditLog.id.desc())
        )
        self.assertEqual(audit.table_code, "rent_confirm")
        self.assertEqual(audit.row_count, query_response.json()["total"])

    def test_staff_tenant_contract_query_does_not_return_sensitive_keys(self) -> None:
        response = self.query(
            "tenant_contract",
            {
                "filters": [{"col": "contact_name", "op": "contains", "val": ""}],
                "sort": [{"col": "room_id", "dir": "asc"}],
                "page": 1,
                "size": 50,
            },
            self.staff,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["total"], 2)
        for row in response.json()["rows"]:
            self.assertNotIn("contact_phone", row)
            self.assertNotIn("rent", row)

    def test_unregistered_column_is_rejected(self) -> None:
        response = self.query(
            "room",
            {
                "filters": [{"col": "not_registered", "op": "eq", "val": "x"}],
                "sort": [],
                "page": 1,
                "size": 50,
            },
        )

        self.assertEqual(response.status_code, 400)

    def test_operator_not_allowed_is_rejected(self) -> None:
        response = self.query(
            "room",
            {
                "filters": [{"col": "room_code", "op": "range", "val": ["A", "Z"]}],
                "sort": [],
                "page": 1,
                "size": 50,
            },
        )

        self.assertEqual(response.status_code, 400)


if __name__ == "__main__":
    unittest.main()
