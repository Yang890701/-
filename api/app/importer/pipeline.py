from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import MetaData, insert, select, text, update
from sqlalchemy.engine import Connection, Engine

SOURCE_FILES = {
    "rent_log": "繳租表單紀錄檔.xlsx",
    "tenant_current": "房客資料表.xlsx",
    "tenant_monthly": "歷程房客資料表.xlsx",
    "mgmt_reminder": "管理提醒專區.xlsx",
}

STAGING_BY_SHEET = {
    "案場總表": "stg_site",
    "電號總表": "stg_meter",
    "房號總表": "stg_room",
    "房號度數紀錄表": "stg_meter_reading",
    "平均電價紀錄表": "stg_avg_price",
    "房客Line資訊": "stg_line_contact",
    "特殊繳租電價設置": "stg_special_price",
    "房號固定費用設定表": "stg_room_fixed_fee",
    "例外款項紀錄表": "stg_exception_charge",
    "繳租確認明細": "stg_rent_confirm",
    "房客資料表": "stg_tenant_contract",
    "歷程房客資料表": "stg_tenant_monthly",
    "管理通知列表": "stg_mgmt_reminder",
}

IMPORT_TABLES = [
    "import_row_error",
    "legacy_key_map",
    "stg_site",
    "stg_meter",
    "stg_room",
    "stg_meter_reading",
    "stg_avg_price",
    "stg_line_contact",
    "stg_special_price",
    "stg_room_fixed_fee",
    "stg_exception_charge",
    "stg_rent_confirm",
    "stg_tenant_contract",
    "stg_tenant_monthly",
    "stg_mgmt_reminder",
    "meter_reading",
    "reading_exception",
    "room_meter_assignment",
    "avg_price",
    "special_price",
    "room_fixed_fee",
    "exception_charge",
    "rent_confirm",
    "tenant_contract",
    "line_contact",
    "mgmt_reminder",
    "room",
    "meter",
    "site",
    "enum_map",
    "enum_management_type",
    "enum_fee_item",
    "enum_meter_category",
    "enum_pay_type",
    "import_batch",
]


ENUM_SEEDS = {
    "management_type": {
        "代管": ("managed", "代管"),
        "一般代管": ("managed", "代管"),
        "投資代管": ("managed", "代管"),
        "代管公司": ("managed_company", "代管公司"),
        "包租": ("master_lease", "包租"),
        "包租公司": ("master_lease_company", "包租公司"),
    },
    "fee_item": {
        "管理費": ("management_fee", "管理費"),
        "雜物處理": ("junk_removal", "雜物處理"),
        "雜物處理費": ("junk_removal", "雜物處理"),
        "雜物代收": ("junk_removal", "雜物處理"),
        "垃圾代收": ("trash_collection", "垃圾代收"),
        "垃圾袋收": ("trash_collection", "垃圾代收"),
        "代收服務": ("collection_service", "代收服務"),
        "水費": ("water_fee", "水費"),
        "網路費": ("internet_fee", "網路費"),
        "清潔費": ("cleaning_fee", "清潔費"),
        "停車費": ("parking_fee", "停車費"),
        "機車車位": ("motorcycle_parking_fee", "機車車位"),
        "房租減免": ("rent_discount", "房租減免"),
        "維護補貼費用": ("maintenance_subsidy", "維護補貼費用"),
        "稅金": ("tax", "稅金"),
    },
    "meter_category": {
        "110V": ("110v", "110V"),
        "220V": ("220v", "220V"),
        "總電表": ("main_meter", "總電表"),
        "子電表": ("sub_meter", "子電表"),
    },
    "pay_type": {
        "補繳": ("supplement", "補繳"),
        "房租預繳-年": ("rent_prepay_year", "房租預繳-年"),
        "房租預繳-半年": ("rent_prepay_half_year", "房租預繳-半年"),
        "租客代墊": ("tenant_advance", "租客代墊"),
        "台電帳單費用": ("taipower_bill", "台電帳單費用"),
    },
}


@dataclass
class TableStats:
    source_rows: int = 0
    imported: int = 0
    excluded: int = 0
    exclude_reasons: Counter[str] = field(default_factory=Counter)
    null_counts: Counter[str] = field(default_factory=Counter)
    duplicate_key_counts: Counter[str] = field(default_factory=Counter)
    fk_orphan_counts: Counter[str] = field(default_factory=Counter)
    money_sums: Counter[str] = field(default_factory=Counter)

    def as_dict(self) -> dict[str, Any]:
        return {
            "source_rows": self.source_rows,
            "imported": self.imported,
            "excluded": self.excluded,
            "exclude_reasons": dict(self.exclude_reasons),
            "null_counts": dict(self.null_counts),
            "duplicate_key_counts": dict(self.duplicate_key_counts),
            "fk_orphan_counts": dict(self.fk_orphan_counts),
            "money_sums": {key: _json_number(value) for key, value in self.money_sums.items()},
        }


def _json_number(value: Any) -> Any:
    if isinstance(value, Decimal):
        return int(value) if value == value.to_integral_value() else float(value)
    return value


class ExcelImporter:
    """Best-effort T4 importer.

    room_meter_assignment is derived from current room.meter_id and observed meter
    reading categories: one open-ended assignment per room/category, effective from
    the earliest observed billing_ym. Exact billing correctness is intentionally
    deferred to later billing validation tasks.
    """

    def __init__(self, engine: Engine, root: Path) -> None:
        self.engine = engine
        self.root = root
        self.meta = MetaData()
        self.stats: dict[str, TableStats] = defaultdict(TableStats)
        self.raw: dict[str, list[dict[str, Any]]] = defaultdict(list)
        self.batch_ids: dict[str, int] = {}
        self.enum_ids: dict[str, dict[str, int]] = defaultdict(dict)
        self.site_ids: dict[str, int] = {}
        self.meter_ids: dict[str, int] = {}
        self.room_ids: dict[str, int] = {}
        self.assignment_ids: dict[tuple[int, str], int] = {}
        self.staging_source_rows: Counter[str] = Counter()
        self.leading_zero_issues: list[dict[str, Any]] = []
        self.roc_anomalies: list[dict[str, Any]] = []

    def run(self) -> dict[str, Any]:
        self.meta.reflect(self.engine)
        with self.engine.begin() as conn:
            self._truncate_import_tables(conn)
            self._create_batches(conn)
            self._load_workbooks(conn)
            self._seed_enums(conn)
            self._insert_sites(conn)
            self._insert_meters(conn)
            self._insert_rooms(conn)
            self._insert_line_contacts(conn)
            self._insert_tenant_contracts(conn)
            self._insert_room_meter_assignments(conn)
            self._insert_meter_readings(conn)
            self._insert_avg_prices(conn)
            self._insert_special_prices(conn)
            self._insert_room_fixed_fees(conn)
            self._insert_exception_charges(conn)
            self._insert_rent_confirms(conn)
            self._insert_mgmt_reminders(conn)
            self._finish_batches(conn)

        return self._report()

    def _truncate_import_tables(self, conn: Connection) -> None:
        names = [name for name in IMPORT_TABLES if name in self.meta.tables]
        quoted = ", ".join(f'"{name}"' for name in names)
        conn.execute(text(f"TRUNCATE TABLE {quoted} RESTART IDENTITY CASCADE"))

    def _create_batches(self, conn: Connection) -> None:
        for key, filename in SOURCE_FILES.items():
            batch_id = conn.execute(
                insert(self.meta.tables["import_batch"])
                .values(source_file=filename, counts={}, checksums={})
                .returning(self.meta.tables["import_batch"].c.id)
            ).scalar_one()
            self.batch_ids[key] = batch_id

    def _load_workbooks(self, conn: Connection) -> None:
        for key, filename in SOURCE_FILES.items():
            path = self.root / filename
            wb = load_workbook(path, read_only=True, data_only=True)
            for ws in wb.worksheets:
                if ws.title == "工作表1":
                    continue
                stg_name = STAGING_BY_SHEET.get(ws.title)
                if not stg_name:
                    continue
                rows = ws.iter_rows(values_only=True)
                headers = next(rows, None)
                if not headers:
                    continue
                clean_headers = [
                    str(h).strip() if h is not None else f"_blank_{i}" for i, h in enumerate(headers)
                ]
                for excel_row_num, values in enumerate(rows, start=2):
                    row = {
                        clean_headers[i]: self._json_value(values[i])
                        for i in range(min(len(clean_headers), len(values)))
                        if clean_headers[i] and not clean_headers[i].startswith("_blank_")
                    }
                    if not any(v not in (None, "") for v in row.values()):
                        continue
                    row["_source_file"] = filename
                    row["_sheet"] = ws.title
                    row["_row"] = excel_row_num
                    self.raw[ws.title].append(row)
                    self.staging_source_rows[ws.title] += 1
                    conn.execute(
                        insert(self.meta.tables[stg_name]).values(
                            batch_id=self.batch_ids[key],
                            source_key=f"{filename}:{ws.title}:{excel_row_num}",
                            raw=row,
                        )
                    )

    def _seed_enums(self, conn: Connection) -> None:
        table_by_domain = {
            "management_type": "enum_management_type",
            "fee_item": "enum_fee_item",
            "meter_category": "enum_meter_category",
            "pay_type": "enum_pay_type",
        }
        for domain, mapping in ENUM_SEEDS.items():
            enum_table = self.meta.tables[table_by_domain[domain]]
            for raw_value, (code, label) in mapping.items():
                if code not in self.enum_ids[domain]:
                    enum_id = conn.execute(
                        insert(enum_table)
                        .values(code=code, label=label, legacy_key=label, source_key="T4 enum seed")
                        .returning(enum_table.c.id)
                    ).scalar_one()
                    self.enum_ids[domain][code] = enum_id
                conn.execute(
                    insert(self.meta.tables["enum_map"]).values(
                        domain=domain,
                        raw_value=raw_value,
                        canonical_value=code,
                    )
                )

    def _insert_sites(self, conn: Connection) -> None:
        seen: set[str] = set()
        for row in self.raw["案場總表"]:
            code = self._clean_text(row.get("案場"))
            stat = self.stats["site"]
            stat.source_rows += 1
            self._count_nulls(stat, row, ["案場", "案場權狀地址"])
            if not code:
                self._row_error(conn, row, "site", "blank site_code")
                continue
            if code in seen:
                stat.duplicate_key_counts["site_code"] += 1
                self._row_error(conn, row, "site", "duplicate site_code")
                continue
            seen.add(code)
            deleted_at = self._deleted_at(row)
            site_id = conn.execute(
                insert(self.meta.tables["site"])
                .values(
                    site_code=code,
                    name=code,
                    address=self._clean_text(row.get("案場權狀地址")),
                    management_unit=self._clean_text(row.get("行政區")),
                    legacy_key=code,
                    source_key=self._source_key(row),
                    deleted_at=deleted_at,
                )
                .returning(self.meta.tables["site"].c.id)
            ).scalar_one()
            self.site_ids[code] = site_id
            self._legacy(conn, "site", code, site_id)
            stat.imported += 1

    def _insert_meters(self, conn: Connection) -> None:
        seen: set[str] = set()
        for row in self.raw["電號總表"]:
            stat = self.stats["meter"]
            stat.source_rows += 1
            code = self._electricity_code(row.get("電號"), row=row, field="電號")
            self._count_nulls(stat, row, ["電號", "戶名"])
            if not code:
                self._row_error(conn, row, "meter", "blank electricity_code")
                continue
            if code in seen:
                stat.duplicate_key_counts["electricity_code"] += 1
                self._row_error(conn, row, "meter", "duplicate electricity_code")
                continue
            seen.add(code)
            meter_id = conn.execute(
                insert(self.meta.tables["meter"])
                .values(
                    electricity_code=code,
                    name=self._clean_text(row.get("戶名")),
                    note=self._clean_text(row.get("備註")),
                    legacy_key=code,
                    source_key=self._source_key(row),
                    deleted_at=self._deleted_at(row),
                )
                .returning(self.meta.tables["meter"].c.id)
            ).scalar_one()
            self.meter_ids[code] = meter_id
            self._legacy(conn, "meter", code, meter_id)
            stat.imported += 1

    def _insert_rooms(self, conn: Connection) -> None:
        seen: set[str] = set()
        for row in self.raw["房號總表"]:
            stat = self.stats["room"]
            stat.source_rows += 1
            room_code = self._clean_text(row.get("房號"))
            site_code = self._clean_text(row.get("案場"))
            meter_code = (
                self._electricity_code(row.get("電號"), row=row, field="電號") if row.get("電號") else None
            )
            self._count_nulls(stat, row, ["房號", "案場", "電號", "管理類型"])
            if not room_code:
                self._row_error(conn, row, "room", "blank room_code")
                continue
            if room_code in seen:
                stat.duplicate_key_counts["room_code"] += 1
                self._row_error(conn, row, "room", "duplicate room_code")
                continue
            site_id = self.site_ids.get(site_code or "")
            if not site_id:
                stat.fk_orphan_counts["site"] += 1
                self._row_error(conn, row, "room", f"orphan site {site_code!r}")
                continue
            meter_id = None
            if meter_code:
                meter_id = self.meter_ids.get(meter_code)
                if not meter_id:
                    # 占位/找不到的電號(如 12345678901) → 房間仍匯入、電號留 NULL，不丟房間
                    stat.fk_orphan_counts["meter_nulled"] += 1
                    meter_id = None
            mgmt_code = self._enum_code("management_type", row.get("管理類型"))
            room_id = conn.execute(
                insert(self.meta.tables["room"])
                .values(
                    site_id=site_id,
                    room_code=room_code,
                    room_name=self._clean_text(row.get("權狀地址")),
                    meter_id=meter_id,
                    management_type_id=self.enum_ids["management_type"].get(mgmt_code or ""),
                    management_type=mgmt_code,
                    management_contact=self._clean_text(row.get("管理單位名稱")),
                    billing_mode=self._clean_text(row.get("電費計算類型")),
                    legacy_key=room_code,
                    source_key=self._source_key(row),
                    deleted_at=self._deleted_at(row),
                )
                .returning(self.meta.tables["room"].c.id)
            ).scalar_one()
            self.room_ids[room_code] = room_id
            self._legacy(conn, "room", room_code, room_id)
            seen.add(room_code)
            stat.imported += 1

    def _insert_line_contacts(self, conn: Connection) -> None:
        seen: set[str] = set()
        for row in self.raw["房客Line資訊"]:
            stat = self.stats["line_contact"]
            stat.source_rows += 1
            line_id = self._clean_text(row.get("Line ID"))
            display = self._clean_text(row.get("Line名稱"))
            self._count_nulls(stat, row, ["Line ID", "Line名稱"])
            if not line_id and not display:
                stat.excluded += 1
                stat.exclude_reasons["blank trailing line row"] += 1
                continue
            if not line_id:
                self._row_error(conn, row, "line_contact", "blank line_id")
                continue
            if line_id in seen:
                stat.duplicate_key_counts["line_id"] += 1
                self._row_error(conn, row, "line_contact", "duplicate line_id")
                continue
            seen.add(line_id)
            line_contact_id = conn.execute(
                insert(self.meta.tables["line_contact"])
                .values(
                    line_id=line_id,
                    display_name=display,
                    first_seen_at=self._dt(row.get("建立日期")),
                    legacy_key=line_id,
                    source_key=self._source_key(row),
                )
                .returning(self.meta.tables["line_contact"].c.id)
            ).scalar_one()
            self._legacy(conn, "line_contact", line_id, line_contact_id)
            stat.imported += 1

    def _insert_tenant_contracts(self, conn: Connection) -> None:
        seen: set[tuple[int, date]] = set()
        for sheet in ["房客資料表", "歷程房客資料表"]:
            for row in self.raw[sheet]:
                stat = self.stats["tenant_contract"]
                stat.source_rows += 1
                room_code = self._clean_text(row.get("房號"))
                self._count_nulls(
                    stat,
                    row,
                    ["房號", "起租年", "起租月", "起租日", "到期年", "到期月", "到期日", "租金", "電話"],
                )
                if row.get("租金") is not None:
                    stat.money_sums["租金"] += self._int(row.get("租金")) or 0
                room_id = self.room_ids.get(room_code or "")
                if not room_id:
                    stat.fk_orphan_counts["room"] += 1
                    self._row_error(conn, row, "tenant_contract", f"orphan room {room_code!r}")
                    continue
                start = self._roc_date(row, "起租")
                # 空起租日 → 以 null 匯入(不丟)；(room,start) 去重，null 視為同鍵(每房至多一筆空起租合約，避免 legacy_key 撞號)
                key = (room_id, start)
                if key in seen:
                    stat.duplicate_key_counts["room_id,lease_start_date"] += 1
                    stat.excluded += 1
                    stat.exclude_reasons["duplicate contract already imported"] += 1
                    continue
                seen.add(key)
                if start is None:
                    stat.null_counts["lease_start_date"] += 1
                start_iso = start.isoformat() if start else "nostart"
                end = self._roc_date(row, "到期")
                phone = self._phone(row.get("電話"), row=row, field="電話")
                contract_id = conn.execute(
                    insert(self.meta.tables["tenant_contract"])
                    .values(
                        room_id=room_id,
                        lease_start_date=start,
                        lease_end_date=end,
                        rent=self._int(row.get("租金")),
                        contact_name=self._clean_text(row.get("承租人")),
                        contact_phone=phone,
                        legacy_key=f"{room_code}_{start_iso}",
                        source_key=self._source_key(row),
                    )
                    .returning(self.meta.tables["tenant_contract"].c.id)
                ).scalar_one()
                self._legacy(conn, "tenant_contract", f"{room_code}_{start_iso}", contract_id)
                stat.imported += 1

    def _insert_room_meter_assignments(self, conn: Connection) -> None:
        earliest: dict[tuple[str, str], str] = {}
        for row in self.raw["房號度數紀錄表"]:
            room_code = self._clean_text(row.get("房號"))
            raw_cat = self._clean_text(row.get("電表類別"))
            category = self._enum_code("meter_category", raw_cat)
            ym = self._ym(row.get("帳單年月"))
            if not room_code or not category or not ym:
                continue
            key = (room_code, category)
            earliest[key] = min(earliest.get(key, ym), ym)
        for (room_code, category), ym in sorted(earliest.items()):
            stat = self.stats["room_meter_assignment"]
            stat.source_rows += 1
            room_id = self.room_ids.get(room_code)
            if not room_id:
                stat.fk_orphan_counts["room"] += 1
                continue
            meter_id = conn.execute(
                select(self.meta.tables["room"].c.meter_id).where(self.meta.tables["room"].c.id == room_id)
            ).scalar_one_or_none()
            if not meter_id:
                stat.fk_orphan_counts["meter"] += 1
                continue
            assignment_id = conn.execute(
                insert(self.meta.tables["room_meter_assignment"])
                .values(
                    room_id=room_id,
                    meter_id=meter_id,
                    effective_from_ym=ym,
                    effective_to_ym=None,
                    meter_category_id=self.enum_ids["meter_category"].get(category),
                    meter_category=category,
                    legacy_key=f"{room_code}_{category}_{ym}",
                    source_key="derived from 房號度數紀錄表",
                )
                .returning(self.meta.tables["room_meter_assignment"].c.id)
            ).scalar_one()
            self.assignment_ids[(room_id, category)] = assignment_id
            self._legacy(conn, "room_meter_assignment", f"{room_code}_{category}_{ym}", assignment_id)
            stat.imported += 1

    def _insert_meter_readings(self, conn: Connection) -> None:
        seen: set[tuple[int, str, str]] = set()
        for row in self.raw["房號度數紀錄表"]:
            stat = self.stats["meter_reading"]
            stat.source_rows += 1
            room_code = self._clean_text(row.get("房號"))
            category = self._enum_code("meter_category", row.get("電表類別"))
            ym = self._ym(row.get("帳單年月"))
            self._count_nulls(stat, row, ["房號", "電號", "帳單年月", "電表類別", "度數"])
            if not room_code or not ym:
                self._row_error(conn, row, "meter_reading", "blank room_code or billing_ym")
                continue
            room_id = self.room_ids.get(room_code)
            if not room_id:
                stat.fk_orphan_counts["room"] += 1
                self._row_error(conn, row, "meter_reading", f"orphan or dirty room {room_code!r}")
                continue
            if not category:
                raw_cat = self._clean_text(row.get("電表類別")) or ""
                if raw_cat.startswith("景平"):
                    self._row_error(
                        conn, row, "meter_reading", "deferred to T16: merged-meter category (景平401-404)"
                    )
                else:
                    self._row_error(
                        conn, row, "meter_reading", f"unmapped meter_category {row.get('電表類別')!r}"
                    )
                continue
            assignment_id = self.assignment_ids.get((room_id, category))
            if not assignment_id:
                self._row_error(conn, row, "meter_reading", "unresolvable room_meter_assignment")
                continue
            key = (assignment_id, ym, "actual")
            if key in seen:
                stat.duplicate_key_counts["assignment_id,billing_ym,reading_kind"] += 1
                self._row_error(conn, row, "meter_reading", "duplicate meter reading natural key")
                continue
            reading = self._int(row.get("度數"))
            if reading is None:
                self._row_error(conn, row, "meter_reading", "blank or invalid reading")
                continue
            seen.add(key)
            reading_id = conn.execute(
                insert(self.meta.tables["meter_reading"])
                .values(
                    assignment_id=assignment_id,
                    billing_ym=ym,
                    reading_kind="actual",
                    reading=reading,
                    legacy_key=f"{room_code}_{category}_{ym}",
                    source_key=self._source_key(row),
                )
                .returning(self.meta.tables["meter_reading"].c.id)
            ).scalar_one()
            self._legacy(conn, "meter_reading", f"{room_code}_{category}_{ym}", reading_id)
            stat.imported += 1

    def _insert_avg_prices(self, conn: Connection) -> None:
        seen: set[tuple[int, str]] = set()
        for row in self.raw["平均電價紀錄表"]:
            stat = self.stats["avg_price"]
            stat.source_rows += 1
            meter_code = self._electricity_code(row.get("電號"), row=row, field="電號")
            ym = self._ym(row.get("帳單年月"))
            meter_id = self.meter_ids.get(meter_code or "")
            if row.get("平均電價") is not None:
                stat.money_sums["平均電價"] += Decimal(str(row.get("平均電價")))
            if not meter_id:
                stat.fk_orphan_counts["meter"] += 1
                self._row_error(conn, row, "avg_price", f"orphan meter {meter_code!r}")
                continue
            key = (meter_id, ym or "")
            if not ym or key in seen:
                stat.duplicate_key_counts["meter_id,billing_ym"] += int(key in seen)
                self._row_error(conn, row, "avg_price", "blank or duplicate billing_ym")
                continue
            seen.add(key)
            avg_id = conn.execute(
                insert(self.meta.tables["avg_price"])
                .values(
                    meter_id=meter_id,
                    billing_ym=ym,
                    price=self._decimal(row.get("平均電價")),
                    legacy_key=self._clean_text(row.get("電號帳單年月")),
                    source_key=self._source_key(row),
                    deleted_at=self._deleted_at(row),
                )
                .returning(self.meta.tables["avg_price"].c.id)
            ).scalar_one()
            self._legacy(
                conn, "avg_price", self._clean_text(row.get("電號帳單年月")) or f"{meter_code}_{ym}", avg_id
            )
            stat.imported += 1

    def _insert_special_prices(self, conn: Connection) -> None:
        for row in self.raw["特殊繳租電價設置"]:
            stat = self.stats["special_price"]
            stat.source_rows += 1
            room_code = self._clean_text(row.get("房號"))
            room_id = self.room_ids.get(room_code or "")
            if not room_id:
                stat.fk_orphan_counts["room"] += 1
                self._row_error(conn, row, "special_price", f"orphan room {room_code!r}")
                continue
            special_id = conn.execute(
                insert(self.meta.tables["special_price"])
                .values(
                    room_id=room_id,
                    meter_category=None,
                    price=self._decimal(row.get("平均電價")) or Decimal(0),
                    legacy_key=room_code,
                    source_key=self._source_key(row),
                )
                .returning(self.meta.tables["special_price"].c.id)
            ).scalar_one()
            self._legacy(conn, "special_price", room_code or str(special_id), special_id)
            stat.imported += 1

    def _insert_room_fixed_fees(self, conn: Connection) -> None:
        seen: set[tuple[int, int]] = set()
        for row in self.raw["房號固定費用設定表"]:
            stat = self.stats["room_fixed_fee"]
            stat.source_rows += 1
            if row.get("金額") is not None:
                stat.money_sums["金額"] += self._int(row.get("金額")) or 0
            room_code = self._clean_text(row.get("房號"))
            room_id = self.room_ids.get(room_code or "")
            item_code = self._enum_code("fee_item", row.get("費用項目"))
            fee_item_id = self.enum_ids["fee_item"].get(item_code or "")
            if not room_id:
                stat.fk_orphan_counts["room"] += 1
                self._row_error(conn, row, "room_fixed_fee", f"orphan room {room_code!r}")
                continue
            if not fee_item_id:
                self._row_error(conn, row, "room_fixed_fee", f"unmapped fee item {row.get('費用項目')!r}")
                continue
            deleted_at = self._deleted_at(row)
            key = (room_id, fee_item_id)
            if not deleted_at and key in seen:
                stat.duplicate_key_counts["room_id,fee_item_id"] += 1
                self._row_error(
                    conn, row, "room_fixed_fee", "duplicate active fixed fee after enum normalization"
                )
                continue
            if not deleted_at:
                seen.add(key)
            fee_id = conn.execute(
                insert(self.meta.tables["room_fixed_fee"])
                .values(
                    room_id=room_id,
                    fee_item_id=fee_item_id,
                    amount=self._int(row.get("金額")) or 0,
                    legacy_key=self._clean_text(row.get("房號費用項目")),
                    source_key=self._source_key(row),
                    deleted_at=deleted_at,
                )
                .returning(self.meta.tables["room_fixed_fee"].c.id)
            ).scalar_one()
            self._legacy(
                conn, "room_fixed_fee", self._clean_text(row.get("房號費用項目")) or str(fee_id), fee_id
            )
            stat.imported += 1

    def _insert_exception_charges(self, conn: Connection) -> None:
        for row in self.raw["例外款項紀錄表"]:
            stat = self.stats["exception_charge"]
            stat.source_rows += 1
            if row.get("金額") is not None:
                stat.money_sums["金額"] += self._int(row.get("金額")) or 0
            room_code = self._clean_text(row.get("房號"))
            room_id = self.room_ids.get(room_code or "")
            if not room_id:
                stat.fk_orphan_counts["room"] += 1
                self._row_error(conn, row, "exception_charge", f"orphan room {room_code!r}")
                continue
            charge_id = conn.execute(
                insert(self.meta.tables["exception_charge"])
                .values(
                    room_id=room_id,
                    billing_ym=self._ym(row.get("帳單年月")),
                    charge_type=self._enum_code("pay_type", row.get("支付類型"))
                    or self._clean_text(row.get("支付類型"))
                    or "exception",
                    amount=self._int(row.get("金額")) or 0,
                    note=self._clean_text(row.get("費用說明")),
                    legacy_key=self._clean_text(row.get("帳單年月_房號_支付類型_費用說明_金額")),
                    source_key=self._source_key(row),
                )
                .returning(self.meta.tables["exception_charge"].c.id)
            ).scalar_one()
            self._legacy(
                conn,
                "exception_charge",
                self._clean_text(row.get("帳單年月_房號_支付類型_費用說明_金額")) or str(charge_id),
                charge_id,
            )
            stat.imported += 1

    def _insert_rent_confirms(self, conn: Connection) -> None:
        versions: Counter[tuple[int, str, str]] = Counter()
        for row in self.raw["繳租確認明細"]:
            stat = self.stats["rent_confirm"]
            stat.source_rows += 1
            for col in ["房租", "電費", "其他費用"]:
                if row.get(col) is not None:
                    stat.money_sums[col] += self._int(row.get(col)) or 0
            room_code = self._clean_text(row.get("房號"))
            room_id = self.room_ids.get(room_code or "")
            ym = self._ym(row.get("帳單年月"))
            if not room_id:
                stat.fk_orphan_counts["room"] += 1
                self._row_error(conn, row, "rent_confirm", f"orphan room {room_code!r}")
                continue
            if not ym:
                # 空帳單年月 → 以 null 匯入(不丟)，這是真實應收帳
                stat.null_counts["billing_ym"] += 1
            key = (room_id, ym, "monthly_receivable")
            versions[key] += 1
            if versions[key] > 1:
                stat.duplicate_key_counts["room_id,billing_ym"] += 1
            rent = self._int(row.get("房租"))
            electricity = self._int(row.get("電費"))
            fixed = self._int(row.get("其他費用"))
            total = sum(v or 0 for v in [rent, electricity, fixed])
            confirm_id = conn.execute(
                insert(self.meta.tables["rent_confirm"])
                .values(
                    room_id=room_id,
                    billing_ym=ym,
                    charge_type="monthly_receivable",
                    run_version=versions[key],
                    status="imported",
                    rent_amount=rent,
                    electricity_amount=electricity,
                    fixed_fee_amount=fixed,
                    exception_amount=None,
                    total_amount=total,
                    amounts={"房租": rent, "電費": electricity, "其他費用": fixed},
                    legacy_key=f"{self._clean_text(row.get('房號帳單年月'))}_{row.get('CDate')}",
                    source_key=self._source_key(row),
                    deleted_at=self._deleted_at(row),
                )
                .returning(self.meta.tables["rent_confirm"].c.id)
            ).scalar_one()
            self._legacy(
                conn,
                "rent_confirm",
                f"{self._clean_text(row.get('房號帳單年月'))}_{row.get('CDate')}",
                confirm_id,
            )
            stat.imported += 1

    def _insert_mgmt_reminders(self, conn: Connection) -> None:
        for row in self.raw["管理通知列表"]:
            stat = self.stats["mgmt_reminder"]
            stat.source_rows += 1
            day = self._int(row.get("日期"))
            content = self._clean_text(row.get("提醒內容"))
            if not day or not content:
                self._row_error(conn, row, "mgmt_reminder", "blank reminder content or day")
                continue
            due = date(2026, 7, min(max(day, 1), 31))
            reminder_id = conn.execute(
                insert(self.meta.tables["mgmt_reminder"])
                .values(
                    target_type="free_text",
                    target_id=None,
                    site_id=None,
                    due_date=due,
                    status=self._clean_text(row.get("頻率")) or "active",
                    created_by=None,
                    legacy_key=content,
                    source_key=self._source_key(row),
                )
                .returning(self.meta.tables["mgmt_reminder"].c.id)
            ).scalar_one()
            self._legacy(conn, "mgmt_reminder", content, reminder_id)
            stat.imported += 1

    def _finish_batches(self, conn: Connection) -> None:
        counts = {name: stats.as_dict() for name, stats in sorted(self.stats.items())}
        for batch_id in self.batch_ids.values():
            conn.execute(
                update(self.meta.tables["import_batch"])
                .where(self.meta.tables["import_batch"].c.id == batch_id)
                .values(counts=counts, finished_at=datetime.now(UTC))
            )

    def _row_error(self, conn: Connection, row: dict[str, Any], table: str, reason: str) -> None:
        stat = self.stats[table]
        stat.excluded += 1
        stat.exclude_reasons[reason] += 1
        batch_key = next(
            (k for k, filename in SOURCE_FILES.items() if filename == row.get("_source_file")), "rent_log"
        )
        conn.execute(
            insert(self.meta.tables["import_row_error"]).values(
                batch_id=self.batch_ids[batch_key],
                source_row=row,
                reason=f"{table}: {reason}",
            )
        )

    def _legacy(self, conn: Connection, domain: str, legacy_key: str | None, new_id: int) -> None:
        if not legacy_key:
            return
        conn.execute(
            insert(self.meta.tables["legacy_key_map"]).values(
                domain=domain, legacy_key=legacy_key, new_id=new_id
            )
        )

    def _report(self) -> dict[str, Any]:
        return {
            "idempotency": "truncate+reload",
            "staging_source_rows": dict(self.staging_source_rows),
            "derivations": {
                "room_meter_assignment": (
                    "Derived best-effort from current room.meter_id plus observed reading categories; "
                    "one open-ended assignment per room/category, effective_from_ym=min(observed billing_ym)."
                ),
                "mgmt_reminder.due_date": "Source has day/frequency only; imported as 2026-07-DD for required DATE column.",
                "avg_price.price": "T4-fix: column is NUMERIC(10,4); decimal 平均電價 preserved (not rounded).",
                "room.meter_id": "T4-fix: rooms with placeholder/unresolved 電號 imported with meter_id=NULL (not dropped).",
                "blank_fields": "T4-fix: rent_confirm blank 帳單年月 and tenant_contract blank 起租日 imported as NULL (not dropped).",
            },
            "tables": {name: stats.as_dict() for name, stats in sorted(self.stats.items())},
            "leading_zero_integrity": self.leading_zero_issues,
            "roc_conversion_anomalies": self.roc_anomalies,
        }

    def _count_nulls(self, stat: TableStats, row: dict[str, Any], keys: list[str]) -> None:
        for key in keys:
            if self._clean_text(row.get(key)) is None:
                stat.null_counts[key] += 1

    def _deleted_at(self, row: dict[str, Any]) -> datetime | None:
        return datetime.now(UTC) if self._int(row.get("DELFLAG")) == 1 else None

    def _enum_code(self, domain: str, value: Any) -> str | None:
        raw = self._clean_text(value)
        if raw is None:
            return None
        mapped = ENUM_SEEDS.get(domain, {}).get(raw)
        return mapped[0] if mapped else None

    def _source_key(self, row: dict[str, Any]) -> str:
        return f"{row.get('_source_file')}:{row.get('_sheet')}:{row.get('_row')}"

    def _roc_date(self, row: dict[str, Any], prefix: str) -> date | None:
        year = self._int(row.get(f"{prefix}年"))
        month = self._int(row.get(f"{prefix}月"))
        day = self._int(row.get(f"{prefix}日"))
        if year is None and month is None and day is None:
            return None
        if year is None or month is None or day is None:
            self.roc_anomalies.append(
                {"source": self._source_key(row), "field": prefix, "reason": "partial date"}
            )
            return None
        if 111 <= year <= 116:
            year += 1911
        elif year < 1911:
            self.roc_anomalies.append(
                {
                    "source": self._source_key(row),
                    "field": prefix,
                    "year": year,
                    "reason": "unexpected ROC year",
                }
            )
            year += 1911
        try:
            return date(year, month, day)
        except ValueError as exc:
            self.roc_anomalies.append({"source": self._source_key(row), "field": prefix, "reason": str(exc)})
            return None

    def _ym(self, value: Any) -> str | None:
        if value is None:
            return None
        if isinstance(value, float):
            value = int(value)
        text_value = str(value).strip()
        if text_value.endswith(".0"):
            text_value = text_value[:-2]
        return text_value if len(text_value) == 6 and text_value.isdigit() else None

    def _electricity_code(self, value: Any, *, row: dict[str, Any], field: str) -> str | None:
        code = self._clean_text(value)
        if code is None:
            return None
        if isinstance(value, int) or (isinstance(value, float) and value.is_integer()):
            padded = str(int(value)).zfill(11)
            if padded != code:
                self.leading_zero_issues.append(
                    {"source": self._source_key(row), "field": field, "raw": code, "repaired": padded}
                )
            return padded
        if code.isdigit() and len(code) < 11:
            padded = code.zfill(11)
            self.leading_zero_issues.append(
                {"source": self._source_key(row), "field": field, "raw": code, "repaired": padded}
            )
            return padded
        return code

    def _phone(self, value: Any, *, row: dict[str, Any], field: str) -> str | None:
        phone = self._clean_text(value)
        if phone is None:
            return None
        if isinstance(value, int) or (isinstance(value, float) and value.is_integer()):
            repaired = str(int(value)).zfill(10)
            self.leading_zero_issues.append(
                {"source": self._source_key(row), "field": field, "raw": phone, "repaired": repaired}
            )
            return repaired
        return phone

    def _clean_text(self, value: Any) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            cleaned = value.strip()
        elif isinstance(value, float) and value.is_integer():
            cleaned = str(int(value))
        else:
            cleaned = str(value).strip()
        return cleaned or None

    def _int(self, value: Any) -> int | None:
        if value is None or value == "":
            return None
        try:
            return int(round(float(value)))
        except (TypeError, ValueError):
            return None

    def _decimal(self, value: Any) -> Decimal | None:
        if value is None or value == "":
            return None
        try:
            return Decimal(str(value))
        except (TypeError, ValueError, InvalidOperation):
            return None

    def _dt(self, value: Any) -> datetime | None:
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day)
        return None

    def _json_value(self, value: Any) -> Any:
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, Decimal):
            return float(value)
        return value
